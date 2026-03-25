"""
attendance_app.py
=================
Run this file to start the application.
  1. GUI opens — pick your Excel file, configure sessions
  2. Click "Generate Attendance Report"
  3. attendance_cleaned.xlsx is written next to your input file
  4. dashboard.py is launched automatically — browser opens at http://localhost:8050

Folder structure (keep both files together):
    your_folder/
        attendance_app.py
        dashboard.py
        your_raw_data.xlsx
        attendance_cleaned.xlsx  ← written here automatically

Install dependencies once:
    pip install pandas openpyxl pillow dash dash-bootstrap-components plotly
"""

import os, sys, subprocess, threading, time
from datetime import datetime, time as dtime

import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ─────────────────────────────────────────────────────────────
# RESOURCE PATH  (works inside a PyInstaller EXE too)
# ─────────────────────────────────────────────────────────────
def resource_path(rel):
    try:    base = sys._MEIPASS
    except: base = os.path.abspath("..")
    return os.path.join(base, rel)

# ─────────────────────────────────────────────────────────────
# TIME HELPERS
# ─────────────────────────────────────────────────────────────
def fmt_12(t):
    if not isinstance(t, dtime): return str(t)
    return datetime.combine(datetime.today(), t).strftime("%I:%M %p")

def add_minutes(t, mins):
    total = max(0, min(t.hour * 60 + t.minute + mins, 23 * 60 + 59))
    return dtime(total // 60, total % 60)

def derive_windows(start, end):
    t_start = datetime.combine(datetime.today(), start)
    t_end   = datetime.combine(datetime.today(), end)
    return {
        "start":         start,
        "end":           end,
        "in_start":      add_minutes(start, -60),
        "in_end":        add_minutes(start, +20),
        "out_start":     end,
        "out_end":       add_minutes(end,   +20),
        "early_out":     add_minutes(end,   -60),
        "total_seconds": max(0, (t_end - t_start).total_seconds()),
    }

# ─────────────────────────────────────────────────────────────
# CLOCK PICKER  (12-hr)
# ─────────────────────────────────────────────────────────────
class ClockPicker(tk.Toplevel):
    def __init__(self, parent, title="Select Time", initial_hour=8, initial_minute=0):
        super().__init__(parent)
        self.title(title); self.resizable(False, False); self.grab_set()
        self.result = None; self._hour24 = initial_hour
        self._minute = initial_minute; self._mode = "hour"
        self.configure(bg="#1E1E2E")
        self._build_ui(); self._draw_clock(); self.wait_window()

    def _hour12(self):
        h = self._hour24 % 12; return h if h else 12

    def _update_display(self):
        self._hour_var.set(f"{self._hour12():02d}")
        self._min_var.set(f"{self._minute:02d}")
        self._ampm_var.set("AM" if self._hour24 < 12 else "PM")

    def _build_ui(self):
        tk.Label(self, text="Select Time", font=("Segoe UI", 13, "bold"),
                 bg="#1E1E2E", fg="#CDD6F4").pack(pady=(14, 4))
        disp = tk.Frame(self, bg="#1E1E2E"); disp.pack()
        self._hour_var = tk.StringVar(value=f"{self._hour12():02d}")
        self._min_var  = tk.StringVar(value=f"{self._minute:02d}")
        self._ampm_var = tk.StringVar(value="AM" if self._hour24 < 12 else "PM")
        self._hour_btn = tk.Label(disp, textvariable=self._hour_var, width=3,
                                   font=("Segoe UI", 32, "bold"), bg="#313244",
                                   fg="#89B4FA", cursor="hand2")
        self._hour_btn.grid(row=0, column=0, padx=4, pady=4, ipady=4)
        self._hour_btn.bind("<Button-1>", lambda e: self._set_mode("hour"))
        tk.Label(disp, text=":", font=("Segoe UI", 32, "bold"),
                 bg="#1E1E2E", fg="#CDD6F4").grid(row=0, column=1)
        self._min_btn = tk.Label(disp, textvariable=self._min_var, width=3,
                                  font=("Segoe UI", 32, "bold"), bg="#313244",
                                  fg="#A6E3A1", cursor="hand2")
        self._min_btn.grid(row=0, column=2, padx=4, pady=4, ipady=4)
        self._min_btn.bind("<Button-1>", lambda e: self._set_mode("minute"))
        self._ampm_btn = tk.Label(disp, textvariable=self._ampm_var, width=3,
                                   font=("Segoe UI", 14, "bold"), bg="#45475A",
                                   fg="#FAB387", cursor="hand2")
        self._ampm_btn.grid(row=0, column=3, padx=(10, 4), ipady=4)
        self._ampm_btn.bind("<Button-1>", lambda e: self._toggle_ampm())
        self.canvas = tk.Canvas(self, width=220, height=220,
                                bg="#1E1E2E", highlightthickness=0)
        self.canvas.pack(padx=20, pady=6)
        self.canvas.bind("<Button-1>",        self._on_click)
        self.canvas.bind("<B1-Motion>",       self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        bf = tk.Frame(self, bg="#1E1E2E"); bf.pack(pady=(4, 14))
        tk.Button(bf, text="Cancel", font=("Segoe UI", 10), bg="#45475A",
                  fg="#CDD6F4", relief="flat", width=8,
                  command=self.destroy).pack(side="left", padx=6)
        tk.Button(bf, text="OK", font=("Segoe UI", 10, "bold"), bg="#89B4FA",
                  fg="#1E1E2E", relief="flat", width=8,
                  command=self._confirm).pack(side="left", padx=6)

    def _toggle_ampm(self):
        self._hour24 = (self._hour24 + 12 if self._hour24 < 12 else self._hour24 - 12)
        self._update_display(); self._draw_clock()

    def _set_mode(self, mode):
        self._mode = mode
        if mode == "hour":
            self._hour_btn.config(bg="#89B4FA", fg="#1E1E2E")
            self._min_btn.config(bg="#313244",  fg="#A6E3A1")
        else:
            self._min_btn.config(bg="#A6E3A1",  fg="#1E1E2E")
            self._hour_btn.config(bg="#313244", fg="#89B4FA")
        self._draw_clock()

    def _draw_clock(self):
        import math
        self.canvas.delete("all")
        cx, cy, r = 110, 110, 90
        self.canvas.create_oval(cx-r, cy-r, cx+r, cy+r,
                                fill="#313244", outline="#585B70", width=2)
        if self._mode == "hour":
            labels  = list(range(1, 13)); current = self._hour12()
        else:
            labels  = [i * 5 for i in range(12)]; current = (self._minute // 5) * 5
        for i, val in enumerate(labels):
            angle = math.radians(i * 30 - 60)
            tx = cx + (r - 18) * math.cos(angle)
            ty = cy + (r - 18) * math.sin(angle)
            is_sel = (val == current)
            if is_sel:
                self.canvas.create_oval(tx-14, ty-14, tx+14, ty+14,
                                        fill="#89B4FA", outline="")
            self.canvas.create_text(tx, ty, text=str(val),
                                    fill="#1E1E2E" if is_sel else "#CDD6F4",
                                    font=("Segoe UI", 9, "bold"))
        hand_angle = (math.radians(self._hour12() * 30 - 90) if self._mode == "hour"
                      else math.radians(self._minute * 6 - 90))
        hx = cx + (r - 34) * math.cos(hand_angle)
        hy = cy + (r - 34) * math.sin(hand_angle)
        self.canvas.create_line(cx, cy, hx, hy, fill="#89B4FA", width=2)
        self.canvas.create_oval(cx-4, cy-4, cx+4, cy+4, fill="#89B4FA", outline="")

    def _angle_to_value(self, x, y):
        import math
        cx, cy = 110, 110
        angle = math.degrees(math.atan2(y - cy, x - cx)) + 90
        if angle < 0: angle += 360
        if self._mode == "hour":
            val = int(round(angle / 30)) % 12; return val if val else 12
        else:
            return (int(round(angle / 6)) % 60 // 5) * 5

    def _on_click(self, e):
        val = self._angle_to_value(e.x, e.y)
        if self._mode == "hour":
            is_pm = self._hour24 >= 12
            self._hour24 = (12 if is_pm else 0) if val == 12 else val + (12 if is_pm else 0)
        else:
            self._minute = val
        self._update_display(); self._draw_clock()

    def _on_drag(self, e):    self._on_click(e)
    def _on_release(self, e): self._set_mode("minute") if self._mode == "hour" else None

    def _confirm(self):
        self.result = dtime(self._hour24, self._minute); self.destroy()


# ─────────────────────────────────────────────────────────────
# SESSION CONFIG DIALOG
# ─────────────────────────────────────────────────────────────
class SessionConfigDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Session Setup"); self.resizable(False, False)
        self.grab_set(); self.configure(bg="#1E1E2E")
        self.result = None
        self.num_sessions = tk.IntVar(value=2)
        self._start_objs = []; self._end_objs = []
        self._start_vars = []; self._end_vars  = []
        self._build_ui(); self.wait_window()

    def _default_times(self, n):
        defaults = []; prev_end_m = 13 * 60
        for i in range(n):
            if i == 0: s_m, e_m = 8 * 60, 13 * 60
            else:      s_m = prev_end_m + 30; e_m = s_m + 300
            s_m = min(s_m, 23 * 60); e_m = min(e_m, 23 * 60 + 59)
            defaults.append({"start": dtime(s_m // 60, s_m % 60),
                              "end":   dtime(e_m // 60, e_m % 60)})
            prev_end_m = e_m
        return defaults

    def _build_ui(self):
        tk.Label(self, text="Session Setup", font=("Segoe UI", 14, "bold"),
                 bg="#1E1E2E", fg="#CDD6F4").pack(pady=(16, 2))
        tk.Label(self, text="Set start and end time for each session",
                 font=("Segoe UI", 9), bg="#1E1E2E", fg="#A6ADC8").pack(pady=(0, 8))
        nf = tk.Frame(self, bg="#1E1E2E"); nf.pack()
        tk.Label(nf, text="Sessions per day:", font=("Segoe UI", 10),
                 bg="#1E1E2E", fg="#CDD6F4").pack(side="left", padx=(0, 10))
        for n in [1, 2, 3, 4]:
            tk.Radiobutton(nf, text=str(n), variable=self.num_sessions, value=n,
                           font=("Segoe UI", 11, "bold"), bg="#1E1E2E", fg="#89B4FA",
                           selectcolor="#313244", activebackground="#1E1E2E",
                           command=self._rebuild).pack(side="left", padx=6)
        outer = tk.Frame(self, bg="#1E1E2E"); outer.pack(padx=20, pady=10, fill="both", expand=True)
        self._canvas = tk.Canvas(outer, bg="#1E1E2E", highlightthickness=0, height=220)
        sb = tk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); self._canvas.pack(side="left", fill="both", expand=True)
        self._inner = tk.Frame(self._canvas, bg="#1E1E2E")
        self._canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>", lambda e:
            self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._rebuild()
        bf = tk.Frame(self, bg="#1E1E2E"); bf.pack(pady=(4, 16))
        tk.Button(bf, text="Cancel", font=("Segoe UI", 10), bg="#45475A",
                  fg="#CDD6F4", relief="flat", width=10,
                  command=self.destroy).pack(side="left", padx=8)
        tk.Button(bf, text="Generate Report", font=("Segoe UI", 10, "bold"),
                  bg="#A6E3A1", fg="#1E1E2E", relief="flat", width=16,
                  command=self._apply).pack(side="left", padx=8)

    def _rebuild(self):
        for w in self._inner.winfo_children(): w.destroy()
        self._start_objs = []; self._end_objs = []
        self._start_vars = []; self._end_vars  = []
        n = self.num_sessions.get(); defaults = self._default_times(n)
        colors = ["#89B4FA", "#A6E3A1", "#FAB387", "#F38BA8"]
        for i in range(n):
            clr = colors[i % len(colors)]; d = defaults[i]
            self._start_objs.append(d["start"]); self._end_objs.append(d["end"])
            row = tk.Frame(self._inner, bg="#2A2A3E",
                           highlightthickness=1, highlightbackground="#585B70")
            row.pack(fill="x", pady=5, padx=4)
            tk.Label(row, text=f"Session {i+1}", font=("Segoe UI", 10, "bold"),
                     bg="#2A2A3E", fg=clr, width=10).pack(side="left", padx=(10, 4), pady=10)
            tk.Label(row, text="Start", font=("Segoe UI", 9),
                     bg="#2A2A3E", fg="#A6ADC8").pack(side="left")
            s_var = tk.StringVar(value=fmt_12(d["start"])); self._start_vars.append(s_var)
            tk.Button(row, textvariable=s_var, font=("Segoe UI", 10, "bold"),
                      bg="#313244", fg=clr, relief="flat", width=10, cursor="hand2",
                      command=lambda si=i: self._pick_start(si)).pack(side="left", padx=6)
            tk.Label(row, text="→", font=("Segoe UI", 12),
                     bg="#2A2A3E", fg="#585B70").pack(side="left")
            tk.Label(row, text="End", font=("Segoe UI", 9),
                     bg="#2A2A3E", fg="#A6ADC8").pack(side="left")
            e_var = tk.StringVar(value=fmt_12(d["end"])); self._end_vars.append(e_var)
            tk.Button(row, textvariable=e_var, font=("Segoe UI", 10, "bold"),
                      bg="#313244", fg=clr, relief="flat", width=10, cursor="hand2",
                      command=lambda si=i: self._pick_end(si)).pack(side="left", padx=6)

    def _pick_start(self, si):
        t0 = self._start_objs[si]
        p  = ClockPicker(self, f"Session {si+1} — Start Time", t0.hour, t0.minute)
        if p.result: self._start_objs[si] = p.result; self._start_vars[si].set(fmt_12(p.result))

    def _pick_end(self, si):
        t0 = self._end_objs[si]
        p  = ClockPicker(self, f"Session {si+1} — End Time", t0.hour, t0.minute)
        if p.result: self._end_objs[si] = p.result; self._end_vars[si].set(fmt_12(p.result))

    def _apply(self):
        self.result = {
            "num_sessions": self.num_sessions.get(),
            "sessions": [derive_windows(self._start_objs[i], self._end_objs[i])
                         for i in range(self.num_sessions.get())],
        }
        self.destroy()


# ─────────────────────────────────────────────────────────────
# CORE PROCESSING  (unchanged logic from your original script)
# ─────────────────────────────────────────────────────────────
def process_and_generate(input_file, cfg):
    num_sessions = cfg["num_sessions"]
    sessions     = cfg["sessions"]
    total_class_seconds = sum(s["total_seconds"] for s in sessions)

    folder      = os.path.dirname(input_file)
    OUTPUT_FILE = os.path.join(folder, "attendance_cleaned.xlsx")

    raw = pd.read_excel(input_file, header=None, skiprows=5)
    current_date = None; records = []

    for _, row in raw.iterrows():
        cell = str(row[0]).strip()
        if "Date" in cell:
            try:    current_date = cell.split(":")[1].strip()
            except: current_date = None
            continue
        if cell == "No." or pd.isna(row[1]): continue
        records.append({"Sr_No": row[0], "Date": current_date,
                         "PNR_Number": row[1], "Name": row[2],
                         "Department": row[3], "Punch_String": row[5]})

    df = pd.DataFrame(records)

    def split_punch(p):
        if pd.isna(p): return []
        return [x.strip() for x in str(p).split(",")]

    df["Punch_List"] = df["Punch_String"].apply(split_punch)

    all_session_cols = []
    for i in range(num_sessions):
        all_session_cols += [f"Session{i+1}_In", f"Session{i+1}_Out"]

    def assign_sessions(time_list):
        ins = [None] * num_sessions; outs = [None] * num_sessions
        for t_str in time_list:
            try:    t = pd.to_datetime(t_str).time()
            except: continue
            for i, s in enumerate(sessions):
                if s["in_start"] <= t < s["out_start"] and ins[i] is None:
                    ins[i] = t; break
                elif s["out_start"] <= t <= s["out_end"] and outs[i] is None:
                    outs[i] = t; break
                elif s["in_end"] < t < s["out_start"] and outs[i] is None:
                    outs[i] = t; break
        result = []
        for i in range(num_sessions):
            result.append(ins[i]); result.append(outs[i])
        return pd.Series(result)

    df[all_session_cols] = df["Punch_List"].apply(assign_sessions)
    df[all_session_cols] = df[all_session_cols].fillna("NA")

    def compute_flags(row):
        flags = {}
        for i, s in enumerate(sessions):
            in_val  = row[f"Session{i+1}_In"]
            out_val = row[f"Session{i+1}_Out"]
            flags[f"Session{i+1}_In_late"]   = isinstance(in_val,  dtime) and in_val  > s["in_end"]
            flags[f"Session{i+1}_Out_early"] = isinstance(out_val, dtime) and out_val < s["early_out"]
        return pd.Series(flags)

    flag_col_names = []
    for i in range(num_sessions):
        flag_col_names += [f"Session{i+1}_In_late", f"Session{i+1}_Out_early"]
    df[flag_col_names] = df.apply(compute_flags, axis=1)

    def session_working_seconds(row, si):
        ic = f"Session{si+1}_In"; oc = f"Session{si+1}_Out"
        if row[ic] != "NA" and row[oc] != "NA":
            t1 = datetime.combine(datetime.today(), row[ic])
            t2 = datetime.combine(datetime.today(), row[oc])
            return max(0, (t2 - t1).total_seconds())
        return 0

    for i in range(num_sessions):
        df[f"Session{i+1}_Seconds"] = df.apply(
            lambda row, si=i: session_working_seconds(row, si), axis=1)

    df["Working_Seconds"] = sum(df[f"Session{i+1}_Seconds"] for i in range(num_sessions))

    for i, s in enumerate(sessions):
        dur = s["total_seconds"]
        df[f"Session{i+1}_PctVal"] = df.apply(
            lambda row, si=i, d=dur: min((row[f"Session{si+1}_Seconds"] / d) * 100, 100.0) if d > 0 else 0.0, axis=1)
        df[f"Session{i+1}_Pct"] = df[f"Session{i+1}_PctVal"].apply(lambda v: f"{v:.2f}%")

    def punch_status(row):
        all_na = all(row[f"Session{i+1}_In"] == "NA" for i in range(num_sessions))
        if all_na: return "Absent"
        for i in range(num_sessions):
            if row[f"Session{i+1}_In"]  == "NA": return f"Absent in Session{i+1}_In"
            if row[f"Session{i+1}_Out"] == "NA": return f"Absent in Session{i+1}_Out"
        if num_sessions > 1:
            if (isinstance(row["Session1_In"], dtime) and isinstance(row["Session1_Out"], dtime)):
                if all(row[f"Session{j+1}_In"] == "NA" and row[f"Session{j+1}_Out"] == "NA"
                       for j in range(1, num_sessions)):
                    return "Partial Absent (Left After S1)"
        has_late  = any(row[f"Session{i+1}_In_late"]   for i in range(num_sessions))
        has_early = any(row[f"Session{i+1}_Out_early"] for i in range(num_sessions))
        if has_late and has_early: return "Present - Late & Left Early"
        if has_late:               return "Present - Late Entry"
        if has_early:              return "Present - Left Early"
        return "Present"

    df["Punch_Status"] = df.apply(punch_status, axis=1)

    for col in all_session_cols:
        df[col] = df[col].apply(lambda v: fmt_12(v) if isinstance(v, dtime) else v)

    # ── Write Excel ──────────────────────────────────────────
    if os.path.exists(OUTPUT_FILE):
        try:    os.remove(OUTPUT_FILE)
        except PermissionError:
            raise PermissionError("Please close 'attendance_cleaned.xlsx' first.")

    writer  = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")
    pd.DataFrame().to_excel(writer, sheet_name="Attendance_Report", index=False)
    ws_att = writer.book["Attendance_Report"]

    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="center")
    date_font    = Font(bold=True, size=12, color="FFFFFF")
    date_fill    = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    hdr_font     = Font(bold=True, color="FFFFFF")
    hdr_fill     = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pct_hdr_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
    orange_fill  = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    alt_fill     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    no_fill      = PatternFill(fill_type=None)

    HEADERS    = ["Sr No", "PNR Number", "Name", "Department"]
    col_widths = [8, 16, 22, 16]
    for i in range(num_sessions):
        HEADERS    += [f"Session{i+1} In", f"Session{i+1} Out", f"Session{i+1} %"]
        col_widths += [15, 15, 12]

    NUM_COLS = len(HEADERS)
    sess_pct_col_indices = {4 + i * 3 + 3 for i in range(num_sessions)}
    current_row = 1

    for date, group in df.groupby("Date"):
        ws_att.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=NUM_COLS)
        dc = ws_att.cell(current_row, 1, f"Date : {date}")
        dc.font = date_font; dc.fill = date_fill; dc.alignment = center; dc.border = border
        ws_att.row_dimensions[current_row].height = 20; current_row += 1

        for ci, h in enumerate(HEADERS, 1):
            c = ws_att.cell(current_row, ci, h)
            c.font = hdr_font
            c.fill = pct_hdr_fill if ci in sess_pct_col_indices else hdr_fill
            c.alignment = center; c.border = border
        ws_att.row_dimensions[current_row].height = 18; current_row += 1

        for rn, (_, r) in enumerate(group.iterrows()):
            sess_vals = []
            for i in range(num_sessions):
                sess_vals += [r[f"Session{i+1}_In"], r[f"Session{i+1}_Out"], r[f"Session{i+1}_Pct"]]
            values = [r["Sr_No"], r["PNR_Number"], r["Name"], r["Department"]] + sess_vals
            row_bg = alt_fill if rn % 2 == 0 else no_fill
            for ci, val in enumerate(values, 1):
                c = ws_att.cell(current_row, ci, val)
                c.alignment = center; c.border = border
                if ci in sess_pct_col_indices:
                    try:    c.fill = orange_fill if float(str(val).replace("%","")) < 75.0 else row_bg
                    except: c.fill = row_bg
                else:
                    c.fill = row_bg
            ws_att.row_dimensions[current_row].height = 16; current_row += 1
        current_row += 1

    for ci, w in enumerate(col_widths, 1):
        ws_att.column_dimensions[get_column_letter(ci)].width = w

    # ── Summary sheet ────────────────────────────────────────
    dates    = sorted(df["Date"].dropna().unique())
    students = df[["PNR_Number", "Name"]].drop_duplicates()
    summary_rows = []

    def status_to_marks(status, n):
        if status.startswith("Present"):                       return ["P"] * n
        if status == "Absent":                                 return ["A"] * n
        if "Partial Absent" in status:                         return ["P"] + ["A"] * (n - 1)
        for i in range(n):
            if f"Absent in Session{i+1}_In"  == status or \
               f"Absent in Session{i+1}_Out" == status:
                return ["A" if j == i else "P" for j in range(n)]
        return ["A"] * n

    for _, student in students.iterrows():
        pnr   = student["PNR_Number"]; name = student["Name"]
        sdata = df[df["PNR_Number"] == pnr]
        row   = [pnr, name]; present = 0; total = 0
        for d in dates:
            day = sdata[sdata["Date"] == d]
            if day.empty:
                marks = ["A"] * num_sessions
            else:
                day_row = day.iloc[0]; marks = []
                for i, s in enumerate(sessions):
                    dur = s["total_seconds"]; attended = day_row[f"Session{i+1}_Seconds"]
                    if dur > 0 and (attended / dur) < 0.75:
                        marks.append("A")
                    else:
                        base = status_to_marks(day_row["Punch_Status"], num_sessions)
                        marks.append(base[i])
            row.extend(marks)
            present += sum(1 for m in marks if m == "P")
            total   += num_sessions
        row.append(f"{present} / {total}")
        summary_rows.append(row)

    columns = ["PNR_Number", "Name"]
    for d in dates:
        for i in range(num_sessions): columns.append(f"{d}_S{i+1}")
    columns.append("Summary")

    summary = pd.DataFrame(summary_rows, columns=columns)
    summary.to_excel(writer, sheet_name="Summary_Report", index=False, header=False)

    ws = writer.book["Summary_Report"]
    ws.insert_rows(1); ws.insert_rows(1)
    col = 3
    for d in dates:
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col + num_sessions - 1)
        ws.cell(1, col).value = d
        for i in range(num_sessions): ws.cell(2, col + i).value = f"S{i+1}"
        col += num_sessions
    ws.merge_cells(start_row=1, start_column=col,  end_row=2, end_column=col)
    ws.merge_cells(start_row=1, start_column=1,    end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2,    end_row=2, end_column=2)
    ws.cell(1, col).value = "Summary"
    ws.cell(1, 1).value   = "PNR_Number"
    ws.cell(1, 2).value   = "Name"

    c2  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rf2 = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    gf2 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    hf2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    df2 = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    af2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    t2  = Side(style="thin"); b2 = Border(left=t2, right=t2, top=t2, bottom=t2)
    tc2 = 2 + len(dates) * num_sessions + 1

    ws.row_dimensions[1].height = 20
    for cell in ws[1]:
        if cell.value: cell.font = Font(bold=True, size=12, color="FFFFFF"); cell.fill = df2
        cell.alignment = c2; cell.border = b2
    ws.row_dimensions[2].height = 18
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hf2
        cell.alignment = c2; cell.border = b2
    for ri, r in enumerate(ws.iter_rows(min_row=3), start=0):
        ws.row_dimensions[ri + 3].height = 16
        rbg = af2 if ri % 2 == 0 else None
        for cell in r:
            cell.alignment = c2; cell.border = b2; v = cell.value
            if v == "A":   cell.fill = rf2
            elif v == "P": cell.fill = gf2
            elif isinstance(v, str) and "/" in v:
                try:
                    pv = int(v.split("/")[0].strip()); tv = int(v.split("/")[1].strip())
                    cell.fill = gf2 if pv == tv else rf2
                except: pass
            elif rbg: cell.fill = rbg

    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(2)].width = 22
    for i in range(len(dates) * num_sessions):
        ws.column_dimensions[get_column_letter(3 + i)].width = 6
    ws.column_dimensions[get_column_letter(tc2)].width = 10
    ws.freeze_panes = "C3"

    writer.close()
    return OUTPUT_FILE


# ─────────────────────────────────────────────────────────────
# LAUNCH DASHBOARD  (called after Excel is written)
# ─────────────────────────────────────────────────────────────
def launch_dashboard(output_excel_path):
    """
    Starts dashboard.py as a separate process.
    Passes the path to attendance_cleaned.xlsx as a command-line argument
    so dashboard.py knows exactly which file to load.
    """
    dashboard_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.py")

    if not os.path.exists(dashboard_script):
        messagebox.showwarning(
            "Dashboard not found",
            f"dashboard.py not found next to this script.\n\n"
            f"Expected location:\n{dashboard_script}\n\n"
            "Report was generated successfully — open dashboard.py manually."
        )
        return

    # Kill any previously running dashboard on port 8050 (best-effort)
    try:
        import psutil
        for proc in psutil.process_iter(["pid", "cmdline"]):
            if proc.info["cmdline"] and "dashboard.py" in " ".join(proc.info["cmdline"]):
                proc.kill()
    except Exception:
        pass  # psutil is optional

    # Launch dashboard.py, passing the Excel file path
    subprocess.Popen(
        [sys.executable, dashboard_script, output_excel_path],
        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
    )


# ─────────────────────────────────────────────────────────────
# MAIN GUI
# ─────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("Attendance Management System")
root.geometry("520x420")
root.resizable(False, False)

# Logo (optional)
if PIL_OK:
    try:
        img  = Image.open(resource_path("attendance_logo.png")).resize((90, 90))
        logo = ImageTk.PhotoImage(img)
        tk.Label(root, image=logo).pack(pady=10)
    except Exception:
        pass

tk.Label(root, text="Attendance Management System",
         font=("Arial", 16, "bold")).pack()
tk.Label(root, text="Automated Attendance Report Generator",
         font=("Arial", 10)).pack(pady=5)

file_path = tk.StringVar()

def choose_file():
    path = filedialog.askopenfilename(
        title="Choose Excel file",
        filetypes=[("Excel Files", "*.xlsx *.xls")])
    if path:
        file_path.set(path)
        file_label.config(text=os.path.basename(path))

def process_file():
    if not file_path.get():
        messagebox.showerror("Error", "Please select an Excel file"); return

    cfg_dialog = SessionConfigDialog(root)
    if cfg_dialog.result is None: return

    try:
        output_path = process_and_generate(file_path.get(), cfg_dialog.result)
    except PermissionError as e:
        messagebox.showerror("File Open", str(e)); return
    except Exception as e:
        messagebox.showerror("Processing Error", str(e)); return

    # ── SUCCESS: launch dashboard then notify user ────────────
    launch_dashboard(output_path)

    messagebox.showinfo(
        "Done!",
        f"Report generated successfully!\n\n"
        f"Saved at:\n{output_path}\n\n"
        f"Dashboard is opening at http://localhost:8050\n"
        f"(may take a few seconds to load)"
    )

tk.Button(root, text="Choose Excel File",
          font=("Arial", 11), command=choose_file).pack(pady=15)

file_label = tk.Label(root, text="No file selected", font=("Arial", 9))
file_label.pack()

tk.Button(root, text="Generate Attendance Report",
          font=("Arial", 12, "bold"), bg="#2E8B57", fg="white",
          command=process_file).pack(pady=30)

root.mainloop()
